import xlrd
from openpyxl import load_workbook
import pandas as pd

#Data Filtering
df = pd.read_excel('data.xlsx')

df = (df.loc[df['Bet Code'] == 131])
df.to_excel('filteredData.xlsx')

#Data Manipulation
path = "filteredData.xlsx"

inputWorkbook = xlrd.open_workbook(path)
inputWorksheet = inputWorkbook.sheet_by_index(0)

count = 0
for cell in inputWorksheet.col(0):
    count += 1

amount = 0
#win = 0
#lose = 0
for cell in range(1, count):
    if (type(inputWorksheet.cell_value(cell,13)) == str):
        break
    if (inputWorksheet.cell_value(cell, 13) >= 2):
        amount += 1000 * inputWorksheet.cell_value(cell, 10) - 1000
        #win += 1
    else:
        amount -= 1000
        #lose += 1

#print(amount)
#print(win)
#print(lose)

#Data Input 
wb = load_workbook("D:/Programming/GitHub/singapore-pools-data-analysis/results.xlsx")
ws = wb["Sheet1"]
wsCell1 = ws.cell(17,1)
wsCell1.value = "Total Goals Over 1.5 (131)"
wsCell2= ws.cell(17,2)
wsCell2.value = amount
wb.save("D:/Programming/GitHub/singapore-pools-data-analysis/results.xlsx")